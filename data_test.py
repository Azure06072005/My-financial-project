import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import os

def create_sample_financial_excel():
    """Create a sample Excel file with Vietnamese financial data"""
    
    # Create quarters
    quarters = ['Q1 2023', 'Q2 2023', 'Q3 2023', 'Q4 2023', 'Q1 2024']
    
    # CDKT Data (Balance Sheet)
    cdkt_data = {
        'Chỉ tiêu': [
            'TÀI SẢN',
            'TÀI SẢN NGẮN HẠN',
            'Tiền và các khoản tương đương',
            'Đầu tư ngắn hạn',
            'Phải thu ngắn hạn',
            'Hàng tồn kho',
            'TÀI SẢN DÀI HẠN',
            'Tài sản cố định hữu hình',
            'Đầu tư dài hạn',
            'NỢ PHẢI TRẢ',
            'Nợ ngắn hạn',
            'Nợ dài hạn',
            'VỐN CHỦ SỞ HỮU',
            'Vốn góp của chủ sở hữu',
            'Lợi nhuận sau thuế chưa phân phối'
        ],
        'Q1 2023': [2850000, 1850000, 450000, 200000, 680000, 520000, 1000000, 800000, 200000, 1200000, 800000, 400000, 1650000, 1000000, 650000],
        'Q2 2023': [2920000, 1920000, 520000, 180000, 720000, 500000, 1000000, 820000, 180000, 1250000, 850000, 400000, 1670000, 1000000, 670000],
        'Q3 2023': [3100000, 2000000, 580000, 220000, 750000, 450000, 1100000, 900000, 200000, 1300000, 900000, 400000, 1800000, 1000000, 800000],
        'Q4 2023': [3250000, 2150000, 650000, 250000, 800000, 450000, 1100000, 920000, 180000, 1350000, 950000, 400000, 1900000, 1000000, 900000],
        'Q1 2024': [3400000, 2250000, 720000, 280000, 850000, 400000, 1150000, 950000, 200000, 1400000, 1000000, 400000, 2000000, 1000000, 1000000]
    }
    
    # KQKD Data (Income Statement)
    kqkd_data = {
        'Chỉ tiêu': [
            'DOANH THU THUẦN',
            'Doanh thu bán hàng',
            'Các khoản giảm trừ',
            'GIÁ VỐN HÀNG BÁN',
            'LỢI NHUẬN GỘP',
            'Chi phí bán hàng',
            'Chi phí quản lý doanh nghiệp',
            'LỢI NHUẬN THUẦN TỪ HOẠT ĐỘNG',
            'Thu nhập khác',
            'Chi phí khác',
            'LỢI NHUẬN TRƯỚC THUẾ',
            'Chi phí thuế thu nhập doanh nghiệp',
            'LỢI NHUẬN SAU THUẾ'
        ],
        'Q1 2023': [1800000, 1850000, 50000, 1200000, 600000, 180000, 150000, 270000, 20000, 15000, 275000, 55000, 220000],
        'Q2 2023': [1950000, 2000000, 50000, 1300000, 650000, 195000, 160000, 295000, 25000, 18000, 302000, 60400, 241600],
        'Q3 2023': [2100000, 2150000, 50000, 1400000, 700000, 210000, 170000, 320000, 30000, 20000, 330000, 66000, 264000],
        'Q4 2023': [2200000, 2250000, 50000, 1450000, 750000, 220000, 180000, 350000, 35000, 22000, 363000, 72600, 290400],
        'Q1 2024': [2300000, 2350000, 50000, 1500000, 800000, 230000, 190000, 380000, 40000, 25000, 395000, 79000, 316000]
    }
    
    # CSTC Data (Financial Ratios)
    cstc_data = {
        'Chỉ tiêu': [
            'CHỈ SỐ THANH TOÁN',
            'Hệ số thanh toán hiện hành',
            'Hệ số thanh toán nhanh',
            'Hệ số thanh toán tiền mặt',
            'CHỈ SỐ HOẠT ĐỘNG',
            'Vòng quay hàng tồn kho',
            'Vòng quay phải thu',
            'Vòng quay tài sản',
            'CHỈ SỐ ĐÒN BẨY',
            'Hệ số nợ trên tài sản',
            'Hệ số nợ trên vốn chủ sở hữu',
            'Hệ số khả năng thanh toán lãi vay',
            'CHỈ SỐ SINH LỜI',
            'ROA (%)',
            'ROE (%)',
            'Tỷ suất lợi nhuận gộp (%)',
            'Tỷ suất lợi nhuận ròng (%)'
        ],
        'Q1 2023': ['', 2.31, 1.66, 0.56, '', 2.31, 2.65, 0.63, '', 0.42, 0.73, 18.33, '', 7.72, 13.33, 33.33, 12.22],
        'Q2 2023': ['', 2.26, 1.67, 0.61, '', 2.60, 2.71, 0.67, '', 0.43, 0.75, 19.67, '', 8.27, 14.47, 33.33, 12.39],
        'Q3 2023': ['', 2.22, 1.72, 0.64, '', 3.11, 2.80, 0.68, '', 0.42, 0.72, 21.33, '', 8.52, 14.67, 33.33, 12.57],
        'Q4 2023': ['', 2.26, 1.79, 0.68, '', 3.22, 2.75, 0.68, '', 0.42, 0.71, 23.53, '', 8.94, 15.28, 34.09, 13.20],
        'Q1 2024': ['', 2.25, 1.85, 0.72, '', 3.75, 2.71, 0.68, '', 0.41, 0.70, 25.33, '', 9.29, 15.80, 34.78, 13.74]
    }
    
    # Create DataFrames
    df_cdkt = pd.DataFrame(cdkt_data).set_index('Chỉ tiêu')
    df_kqkd = pd.DataFrame(kqkd_data).set_index('Chỉ tiêu')
    df_cstc = pd.DataFrame(cstc_data).set_index('Chỉ tiêu')
    
    # Create Excel workbook
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # Add CDKT sheet
    ws_cdkt = wb.create_sheet('CDKT')
    # Add 5 empty rows (data starts at row 6)
    for i in range(5):
        ws_cdkt.append([''] * 6)
    # Add headers and data
    ws_cdkt.append(['Chỉ tiêu'] + list(df_cdkt.columns))
    for row in dataframe_to_rows(df_cdkt, index=True, header=False):
        ws_cdkt.append(row)
    
    # Add KQKD sheet
    ws_kqkd = wb.create_sheet('KQKD')
    # Add 5 empty rows
    for i in range(5):
        ws_kqkd.append([''] * 6)
    # Add headers and data
    ws_kqkd.append(['Chỉ tiêu'] + list(df_kqkd.columns))
    for row in dataframe_to_rows(df_kqkd, index=True, header=False):
        ws_kqkd.append(row)
    
    # Add CSTC sheet
    ws_cstc = wb.create_sheet('CSTC')
    # Add 5 empty rows
    for i in range(5):
        ws_cstc.append([''] * 6)
    # Add headers and data
    ws_cstc.append(['Chỉ tiêu'] + list(df_cstc.columns))
    for row in dataframe_to_rows(df_cstc, index=True, header=False):
        ws_cstc.append(row)
    
    # Save the file
    filename = 'HPG_hn_Q2_2025.xlsx'
    wb.save(filename)
    
    print(f"✅ Sample Excel file created: {filename}")
    print(f"📁 Location: {os.path.abspath(filename)}")
    print(f"📊 Sheets included: CDKT, KQKD, CSTC")
    print(f"💡 This file is ready to test with your Flask application!")
    
    # Show file info
    file_size = os.path.getsize(filename) / 1024  # KB
    print(f"📏 File size: {file_size:.1f} KB")
    
    return filename

def create_simple_version():
    """Create a simpler version with just essential data"""
    
    # Create a simpler structure
    simple_data = {
        'CDKT': {
            'Chỉ tiêu': ['Tài sản', 'Nợ phải trả', 'Vốn chủ sở hữu'],
            'Q1 2024': [1000000, 400000, 600000],
            'Q2 2024': [1200000, 450000, 750000]
        },
        'KQKD': {
            'Chỉ tiêu': ['Doanh thu', 'Chi phí', 'Lợi nhuận'],
            'Q1 2024': [500000, 350000, 150000],
            'Q2 2024': [600000, 400000, 200000]
        },
        'CSTC': {
            'Chỉ tiêu': ['ROA (%)', 'ROE (%)', 'Tỷ suất lợi nhuận (%)'],
            'Q1 2024': [15.0, 25.0, 30.0],
            'Q2 2024': [16.7, 26.7, 33.3]
        }
    }
    
    with pd.ExcelWriter('Simple_Financial_Data.xlsx', engine='openpyxl') as writer:
        for sheet_name, data in simple_data.items():
            df = pd.DataFrame(data).set_index('Chỉ tiêu')
            df.to_excel(writer, sheet_name=sheet_name, startrow=5)
    
    print("✅ Simple Excel file created: Simple_Financial_Data.xlsx")

if __name__ == "__main__":
    print("🏗️  Creating sample financial Excel files...\n")
    
    # Create full version
    create_sample_financial_excel()
    print()
    
    # Create simple version
    create_simple_version()
    
    print("\n🎯 Instructions:")
    print("1. Use either file to test your Flask application")
    print("2. Upload the file through your React frontend")
    print("3. The data should process correctly and display in tables")
    print("\n💡 Both files follow the exact format expected by your dataframe.py logic!")